import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import create_engine

import datetime
from datetime import date
today_table_name = datetime.date.today().strftime('%d_%m_%Y')

def export_mysql_to_formatted_excel():
    # ==========================================
    # 1. DATABASE & FILE CONFIGURATION
    # ==========================================
    DB_USER = "root"
    DB_PASSWORD = "admin"  # Replace with your actual password
    DB_HOST = "localhost"
    DB_NAME = "school"  # Replace with your actual database name
    TARGET_TABLE = f"{today_table_name}"
    OUTPUT_FILE = f"{today_table_name}.xlsx"
    SHEET_NAME = "Attendance_Data"

    try:
        # ==========================================
        # 2. EXTRACT DATA VIA SQLALCHEMY ENGINE
        # ==========================================
        print("Connecting to the MySQL database...")
        # Creates a standard connection string URI for SQLAlchemy
        connection_url = (
            f"mysql+mysqlconnector://{DB_USER}:{DB_PASSWORD}@{DB_HOST}/{DB_NAME}"
        )
        engine = create_engine(connection_url)

        query = f"SELECT * FROM `{TARGET_TABLE}`"

        print(f"Fetching data from table '{TARGET_TABLE}'...")
        # Using the engine fixes the UserWarning error cleanly
        df = pd.read_sql(query, con=engine)
        print("MySQL data successfully fetched.")

        # ==========================================
        # 3. INITIAL EXPORT VIA PANDAS
        # ==========================================
        print(f"Writing raw data to '{OUTPUT_FILE}'...")
        with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

        # ==========================================
        # 4. APPLY EXCEL TABLE FORMATTING & AUTO-FIT
        # ==========================================
        print("Applying Excel Table layouts and auto-fitting columns...")
        wb = load_workbook(OUTPUT_FILE)
        ws = wb[SHEET_NAME]

        # Determine structural boundaries dynamically
        min_col = ws.min_column
        max_col = ws.max_column
        min_row = ws.min_row
        max_row = ws.max_row

        # Create coordinate range (e.g., "A1:D3")
        data_range = f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}"

        # Initialize native Excel Table object (Name must not contain spaces)
        excel_table = Table(displayName="AttendanceTable", ref=data_range)

        # Apply a clean, structured table theme (Medium 9 is the standard blue style)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        excel_table.tableStyleInfo = style
        ws.add_table(excel_table)

        # Auto-fit columns safely across all openpyxl versions
        # Iterating over col_idx addresses the 'tuple' object attribute crash
        for col_idx in range(min_col, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)

            # Look through every row inside this specific column index
            for row_idx in range(min_row, max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))

            # +5 padding keeps filter arrows from blocking header text labels
            ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

        # Save all layout transformations
        wb.save(OUTPUT_FILE)
        print(f"Process complete! Final formatted spreadsheet saved to: {OUTPUT_FILE}")

    except Exception as e:
        print(f"An unexpected error occurred: {e}")


if __name__ == "__main__":
    export_mysql_to_formatted_excel()
