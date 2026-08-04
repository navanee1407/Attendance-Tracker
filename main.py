import os
import barcode
from barcode.writer import ImageWriter
import mysql.connector
import datetime
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import create_engine

def Generate_add():
    
    mydb = mysql.connector.connect(host='localhost',user='root',password='admin',database='school')
    mycursor = mydb.cursor()

    create_table_query='create table if not exists Student_details(rollNumber int(5) primary key, studentName varchar(50))'
    mycursor.execute(create_table_query)

    rollNumber = input("Enter the student roll number: ")
    studentName = input('Enter the student name:')

    # Get the folder where this script is located
    script_folder = os.path.dirname(os.path.abspath(__file__))

    barcode_class = barcode.get_barcode_class('code128')
    my_barcode = barcode_class(rollNumber, writer=ImageWriter())

    filename = os.path.join(script_folder, f"student_{rollNumber}")
    saved_path = my_barcode.save(filename)

    print(f"Barcode successfully saved as: {saved_path}")

    insert_query ='insert into Student_Details values (%s,%s)'
    insert_values= (rollNumber,studentName)
    mycursor.execute(insert_query,insert_values)
    mydb.commit()

def Marker():
    
    mydb=mysql.connector.connect(host='localhost',user='root',password='admin',database='school')
    mycursor=mydb.cursor()

    today=datetime.datetime.now().strftime('%d/%m/%Y-%H:%M')
    today_table_name = datetime.date.today().strftime('%d_%m_%Y')


    def create_today_table():
        create_today_table_query = f""" 
        create table if not exists `{today_table_name}`(
            rollnumber int(5) primary key,
            studentName varchar(50),
            Attendance_time varchar(20),
            Status varchar(10)
        )
        """
        mycursor.execute(create_today_table_query)

    def mark_present():
        select_studentDetails_query='select * from student_details'
        mycursor.execute(select_studentDetails_query)
        data=mycursor.fetchall()

        rollNumber=int(input('Scan the barcode:'))
        
        select_present_data_query=f'select * from {today_table_name} where rollnumber=%s'
        mycursor.execute(select_present_data_query,(rollNumber,))
        present_data=mycursor.fetchone()

        if present_data is None :
            for i in data:
                if i[0] == rollNumber:
                    studentName=i[1]
                    Status = 'Present'
                    

            insert_query = f'insert ignore into `{today_table_name}` values (%s,%s,%s,%s)'
            insert_values = (rollNumber,studentName,today,Status)

            mycursor.execute(insert_query,insert_values)
            mydb.commit()

            print(f'Attendance marked for {studentName}')

        elif present_data[3]=='Absent':
            change=input('Student marked Absent does changes required?(y/n)')
            if change.lower()=='y':
                update_query=f"update {today_table_name} set status='Present',Attendance_time=%s where rollnumber=%s"
                mycursor.execute(update_query,(today,rollNumber))
                mydb.commit()
                print('Changes made successfully')
            else:
                pass
    def mark_absent():
        select_absent_query = f'select A.rollNumber,A.studentName,B.Attendance_time,B.Status from student_details A left join {today_table_name} B on A.rollNumber=B.rollNumber'
        mycursor.execute(select_absent_query)
        absent_data=mycursor.fetchall()

        for i in absent_data:
            if i[3] == None:
                Status = 'Absent'
                rollNumber=i[0]
                studentName=i[1]

        insert_absent_query = f'insert into {today_table_name} (rollNumber,studentName,Status) values (%s,%s,%s)'
        insert_absent_values = (rollNumber,studentName,Status)

        mycursor.execute(insert_absent_query,insert_absent_values)
        mydb.commit()
        print('Attendance closed successfully')
        
        

    create_today_table()

    print("'1' for Mark Attendance")
    print("'2' for Close Attendance")
    cont='y'
    while (cont.lower()=='y'):
        
        opt=int(input('Choose your option:'))
        if opt==1:
            mark_present()
        elif opt==2:
            mark_absent()
        cont=input('Do you want to continue(y/n:)')

def Export():
    
    today_table_name = datetime.date.today().strftime('%d_%m_%Y')

    def export_mysql_to_formatted_excel():
        # ==========================================
        # 1. DATABASE & FILE CONFIGURATION
        # ==========================================
        DB_USER = "root"
        DB_PASSWORD = "admin"  # Replace with your actual password
        DB_HOST = "localhost"
        DB_NAME = "school"  # Replace with your actual database name
        TARGET_TABLE = f"{today_table_name}"
        OUTPUT_FILE = f"{today_table_name}.xlsx"
        SHEET_NAME = "Attendance_Data"

        try:
            # ==========================================
            # 2. EXTRACT DATA VIA SQLALCHEMY ENGINE
            # ==========================================
            print("Connecting to the MySQL database...")
            # Creates a standard connection string URI for SQLAlchemy
            connection_url = (
                f"mysql+mysqlconnector://{DB_USER}:{DB_PASSWORD}@{DB_HOST}/{DB_NAME}"
            )
            engine = create_engine(connection_url)

            query = f"SELECT * FROM `{TARGET_TABLE}`"

            print(f"Fetching data from table '{TARGET_TABLE}'...")
            # Using the engine fixes the UserWarning error cleanly
            df = pd.read_sql(query, con=engine)
            print("MySQL data successfully fetched.")

            # ==========================================
            # 3. INITIAL EXPORT VIA PANDAS
            # ==========================================
            print(f"Writing raw data to '{OUTPUT_FILE}'...")
            with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

            # ==========================================
            # 4. APPLY EXCEL TABLE FORMATTING & AUTO-FIT
            # ==========================================
            print("Applying Excel Table layouts and auto-fitting columns...")
            wb = load_workbook(OUTPUT_FILE)
            ws = wb[SHEET_NAME]

            # Determine structural boundaries dynamically
            min_col = ws.min_column
            max_col = ws.max_column
            min_row = ws.min_row
            max_row = ws.max_row

            # Create coordinate range (e.g., "A1:D3")
            data_range = f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}"

            # Initialize native Excel Table object (Name must not contain spaces)
            excel_table = Table(displayName="AttendanceTable", ref=data_range)

            # Apply a clean, structured table theme (Medium 9 is the standard blue style)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            excel_table.tableStyleInfo = style
            ws.add_table(excel_table)

            # Auto-fit columns safely across all openpyxl versions
            # Iterating over col_idx addresses the 'tuple' object attribute crash
            for col_idx in range(min_col, max_col + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)

                # Look through every row inside this specific column index
                for row_idx in range(min_row, max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        max_len = max(max_len, len(str(cell_value)))

                # +5 padding keeps filter arrows from blocking header text labels
                ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

            # Save all layout transformations
            wb.save(OUTPUT_FILE)
            print(f"Process complete! Final formatted spreadsheet saved to: {OUTPUT_FILE}")

        except Exception as e:
            print(f"An unexpected error occurred: {e}")


    if __name__ == "__main__":
        export_mysql_to_formatted_excel()

